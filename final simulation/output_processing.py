# read the number of conflicts, number of lane-changing and simulation time from simulation output
# then write them into an excel file

import os
from openpyxl import load_workbook
import re

source_dir = "C:\\Users\\shaw1\\Desktop\\sumo\\hete\\results"
file_home = 'C:\\Users\\shaw1\\Desktop\\1.xlsx'

col = 0
count = 0
time = 0
# count the number of conflicts
with open(os.path.join(source_dir, 'output_ssm.xml'), 'r') as ff:
	header_text = ff.readlines()
	for i in header_text:
		if 'myroutes2' in i:
			col += 1

		if 'type="3"' in i:
			count += 1
print('number of conflicts=' + str(count))

# count the number of lane-changing
lc_number = 0
with open(os.path.join(source_dir, 'LG.xml'), 'r') as ff:
	header_text1 = ff.readlines()
	lc_number = len(header_text1) - 60
	print('number of lane chang= ' + str(lc_number))

# read the simulation time

with open(os.path.join(source_dir, 'full_output_sumo.xml'), 'r') as ff:
	lines = ff.readlines()
	for i in reversed(lines):
		if 'timestep' in i:
			time = float(re.findall(r"\d+\.?\d*", i)[0])
			break


wb = load_workbook(filename=file_home)
sheet_ranges = wb['Sheet2']
ws = wb['Sheet2']

if col == 1:
	lc_name = 'A'
	time_name = 'B'
	col_name = 'C'
	print('IDM')
else:
	lc_name = 'D'
	time_name = 'E'
	col_name = 'F'
	print('CACC')


for i in range(2, 52):
	name = col_name + str(i)
	name_lc = lc_name + str(i)
	name_time = time_name + str(i)
	if sheet_ranges[name].value is None:
		ws[name] = count
		ws[name_lc] = lc_number
		ws[name_time] = time
		print(i)
		break

wb.save(file_home)

